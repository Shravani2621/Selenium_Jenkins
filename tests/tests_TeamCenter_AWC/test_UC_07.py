# Revise an Item
import time
from selenium import webdriver
import pytest
from selenium.webdriver.common.by import By
from pynput.keyboard import Key, Controller
import openpyxl
from selenium.webdriver import Chrome


# @pytest.mark.login
def test():
    driver = Chrome()
    domain = "http://inbr2wks-249497:3000/"
    driver.get(domain)
    driver.set_window_size(1500, 900)
    time.sleep(2)

    path = r"C:\Users\sss927832\Selenium+TC\Selenium+TC\Book1.xlsx"
    b = openpyxl.load_workbook(path)
    sheet = b.active
    c1 = sheet.cell(row=2, column=5)
    c2 = sheet.cell(row=2, column=7)
    driver.find_element(By.NAME, "username").send_keys(c1.value)
    driver.find_element(By.NAME, "password").send_keys(c1.value)
    time.sleep(5)

    # driver.find_element(By.NAME, "username").send_keys("infodba")
    # driver.find_element(By.NAME, "password").send_keys("infodba")
    # time.sleep(5)
    driver.find_element(By.CSS_SELECTOR, ".sw-button.accent-caution").click()
    #time.sleep(10)
    # driver.find_element(By.CSS_SELECTOR,
    #                     ".sw-column.aw-tile-tileContainer.aw-theme-locationsTile.aw-tile-smallSize").click()
    # time.sleep(5)
    # driver.find_element(By.CSS_SELECTOR, ".sw-aria-border:nth-child(10)").click()
    time.sleep(20)
    driver.find_element(By.CSS_SELECTOR, ".aw-uiwidgets-searchBox").click()
    time.sleep(15)
    driver.find_element(By.CSS_SELECTOR, ".aw-search-globalSearchLinksPanel1").click()
    time.sleep(15)
    driver.find_element(By.CSS_SELECTOR, ".aw-panel-header svg").click()
    time.sleep(15)
    driver.find_element(By.XPATH, "/html[1]/body[1]/div[1]/div[1]/div[2]/div[1]/div[1]/div[1]/div[1]/div[1]/div["
                                  "2]/div[1]/div[1]/div[1]/aside[1]/div[2]/form[1]/div[2]/div[1]/div[1]/div[1]/label["
                                  "1]/div[1]/input[1]").clear()
    driver.find_element(By.XPATH, "/html[1]/body[1]/div[1]/div[1]/div[2]/div[1]/div[1]/div[1]/div[1]/div[1]/div["
                                  "2]/div[1]/div[1]/div[1]/aside[1]/div[2]/form[1]/div[2]/div[1]/div[1]/div[1]/label["
                                  "1]/div[1]/input[1]").send_keys("Item..")
    # driver.find_element(By.CSS_SELECTOR, "li:nth-child(6) > .sw-aria-border > .sw-row").click()
    time.sleep(5)
    keyboard = Controller()
    keyboard.press(Key.enter)
    keyboard.release(Key.enter)
    time.sleep(10)
    driver.find_element(By.CSS_SELECTOR, ".sw-property:nth-child(2) > .sw-property-val").click()
    time.sleep(5)
    driver.find_element(By.CSS_SELECTOR, ".sw-property:nth-child(2) > .sw-property-val").send_keys(c2.value)
    time.sleep(5)
    driver.find_element(By.CSS_SELECTOR, ".sw-button > div").click()
    time.sleep(10)
    driver.find_element(By.CSS_SELECTOR, "#ObjectSet_1_Provider_row2_col2").click()
    time.sleep(5)
    driver.find_element(By.CSS_SELECTOR, ".aw-commandId-Awp0NewGroup").click()
    time.sleep(5)
    driver.find_element(By.CSS_SELECTOR, ".aw-list-command:nth-child(3)").click()
    time.sleep(5)
    driver.find_element(By.CSS_SELECTOR, ".sw-button").click()
    time.sleep(10)
