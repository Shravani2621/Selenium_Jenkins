import time
import pytest
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from pynput.keyboard import Key, Controller
from webdriver_manager.microsoft import EdgeChromiumDriverManager

@pytest.fixture
def setup():
    options = webdriver.EdgeOptions()
    driver = webdriver.Edge(EdgeChromiumDriverManager().install(), options=options)

    #driver = webdriver.Edge(EdgeChromiumDriverManager().install(), options=webdriver.EdgeOptions())
    domain = "http://inbr2wks-249497:3000/"
    driver.get(domain)
    time.sleep(2)
    yield driver
    driver.quit()

def test_create_item(setup):
    driver = setup
    path = r"C:\Users\sss927832\OneDrive - Tata Technologies\Desktop\Book1.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    username_cell = sheet.cell(row=5, column=1)
    password_cell = sheet.cell(row=6, column=1)

    # Login
    driver.find_element(By.NAME, "username").send_keys(username_cell.value)
    driver.find_element(By.NAME, "password").send_keys(password_cell.value)
    time.sleep(5)
    driver.find_element(By.CSS_SELECTOR, ".sw-button.accent-caution").click()
    time.sleep(10)

    # Navigate to create item page
    driver.find_element(By.CSS_SELECTOR, ".sw-column.aw-tile-tileContainer.aw-theme-locationsTile.aw-tile-smallSize").click()
    time.sleep(5)
    driver.find_element(By.CSS_SELECTOR, ".aw-commands-commandIconButton.aw-commands-command.aw-commandId-Awp0ShowCreateObject.aw-commands-commandWrapperHorizontal").click()
    time.sleep(5)

    # Fill item details and submit
    driver.find_element(By.XPATH, "//div[@class='sw-lov-container aw-widgets-droppable']/child::input").send_keys("Item")
    time.sleep(2)
    keyboard = Controller()
    keyboard.press(Key.enter)
    keyboard.release(Key.enter)
    time.sleep(2)
    driver.find_element(By.NAME, "object_name").send_keys("Item2")
    time.sleep(2)
    driver.find_element(By.CSS_SELECTOR, ".aw-panel-footer.sw-row.flex-wrap > .sw-button").click()
    time.sleep(5)
