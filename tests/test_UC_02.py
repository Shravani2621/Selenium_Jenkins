# Creating a Part
import time
from selenium import webdriver
import pytest
from selenium.webdriver.common.by import By
from pynput.keyboard import Key, Controller
import openpyxl
from selenium.webdriver import Chrome
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

def wait_for_window(self, timeout=2):
    time.sleep(round(timeout / 1000))
    wh_now = self.driver.window_handles
    wh_then = self.vars["window_handles"]
    if len(wh_now) > len(wh_then):
        return set(wh_now).difference(set(wh_then)).pop()


# @pytest.mark.login
def test():
    driver = Chrome()
    domain = "http://inbr2wks-249497:3000/"
    driver.get(domain)
    driver.set_window_size(1100, 600)
    time.sleep(2)

    path = r"C:\Users\sss927832\Selenium+TC\Selenium+TC\Book1.xlsx"
    b = openpyxl.load_workbook(path)
    sheet = b.active
    c1 = sheet.cell(row=2, column=5)
    c2 = sheet.cell(row=2, column=8)
    driver.find_element(By.NAME, "username").send_keys(c1.value)
    driver.find_element(By.NAME, "password").send_keys(c1.value)
    time.sleep(5)

    # driver.find_element(By.NAME, "username").send_keys("infodba")
    # driver.find_element(By.NAME, "password").send_keys("infodba")
    # time.sleep(5)
    driver.find_element(By.CSS_SELECTOR, ".sw-button.accent-caution").click()
    time.sleep(40)
    driver.find_element(By.CSS_SELECTOR,
                        ".sw-column.aw-tile-tileContainer.aw-theme-locationsTile.aw-tile-smallSize").click()
    time.sleep(5)
    driver.find_element(By.CSS_SELECTOR, ".aw-commands-commandIconButton.aw-commands-command.aw-commandId"
                                         "-Awp0ShowCreateObject.aw-commands-commandWrapperHorizontal").click()
    time.sleep(5)
    # driver.find_element(By.XPATH, "//a[contains(@aria-label, 'Part')]")
    # driver.find_element(By.CSS_SELECTOR, "li:nth-child(3) > .sw-aria-border > .sw-row").click()
    driver.find_element(By.CSS_SELECTOR, ".aw-widgets-droppable > .sw-aria-border").click()
    time.sleep(5)
    driver.find_element(By.XPATH, "//div[@class='sw-lov-container aw-widgets-droppable']/child::input").send_keys(
        "Part")
    keyboard = Controller()
    keyboard.press(Key.enter)
    keyboard.release(Key.enter)
    wait = WebDriverWait(driver, 10)  # 10 seconds timeout
    element = wait.until(EC.element_to_be_clickable((By.NAME, "object_name")))
    element.click()
    time.sleep(3)
    driver.find_element(By.NAME, "object_name").send_keys(c2.value)
    time.sleep(15)
    driver.find_element(By.CSS_SELECTOR, ".aw-panel-footer.sw-row.flex-wrap > .sw-button").click()
    time.sleep(5)
